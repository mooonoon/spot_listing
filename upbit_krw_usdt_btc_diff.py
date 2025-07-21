import requests
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os


def get_upbit_markets():
    """获取Upbit交易所的所有市场信息"""
    url = "https://api.upbit.com/v1/market/all"
    headers = {"accept": "application/json"}
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"请求出错: {e}")
        return []


def get_coin_listing_dates():
    """尝试获取加密货币的上币日期（Upbit未直接提供，此为示例实现）"""
    # 实际应用中应替换为真实的上币日期数据
    # 这里使用简化的示例，按字母顺序模拟上币日期
    return {}


def filter_pairs(markets, prefix):
    """从市场信息中筛选出特定前缀的交易对"""
    return [market for market in markets if market['market'].startswith(prefix)]


def sort_by_listing_date(pairs, listing_dates=None):
    """按上币日期对交易对进行排序"""
    if not listing_dates:
        # 如果没有上币日期数据，按交易对名称排序
        return sorted(pairs, key=lambda x: x['market'])

    # 使用上币日期排序，未知日期的交易对放在最后
    return sorted(pairs,
                  key=lambda x: listing_dates.get(x['market'].split('-')[1], '9999-99-99'))


def save_to_excel(pairs, sheet_name, writer, base_currency=None):
    """将交易对信息保存到Excel的指定工作表中，支持空数据"""
    workbook = writer.book

    # 如果没有数据，创建空的DataFrame
    if not pairs:
        df = pd.DataFrame(columns=[
            '交易对代码', '基础货币', '报价货币',
            '韩文名称', '英文名称', '市场警告', '上币日期(近似)'
        ])
        if base_currency:
            sheet_name = f'only_{base_currency}_pairs'
            df['基础货币'] = base_currency
    else:
        df = pd.DataFrame(pairs)
        df = df.rename(columns={
            'market': '交易对代码',
            'korean_name': '韩文名称',
            'english_name': '英文名称'
        })

        # 添加基础货币和报价货币列
        if 'only_' in sheet_name:
            base_currency = sheet_name.split('_')[1]
            df['基础货币'] = base_currency
        else:
            df['基础货币'] = sheet_name.split('_')[0].upper()

        df['报价货币'] = df['交易对代码'].apply(lambda x: x.split('-')[1])

        # 添加模拟的上币日期列（实际应用中应替换为真实数据）
        listing_dates = get_coin_listing_dates()
        df['上币日期(近似)'] = df['报价货币'].apply(lambda x: listing_dates.get(x, '未知'))

    # 确保所有列存在
    required_columns = ['交易对代码', '基础货币', '报价货币', '韩文名称', '英文名称', '上币日期(近似)']
    if 'market_warning' in df.columns:
        required_columns.append('market_warning')
        df = df.rename(columns={'market_warning': '市场警告'})
    else:
        df['市场警告'] = ''
        required_columns.append('市场警告')

    df = df[required_columns]

    # 写入Excel
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = workbook[sheet_name]

    # 设置列宽
    for i, col in enumerate(df.columns):
        column_width = max(len(str(x)) for x in df[col]) if not df.empty else len(col)
        column_width = max(column_width, len(col)) + 2
        col_letter = get_column_letter(i + 1)
        worksheet.column_dimensions[col_letter].width = column_width

    # 设置表头样式
    if not df.empty:
        header_font = Font(bold=True)
        header_alignment = Alignment(vertical='top', wrap_text=True)
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')

        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1, value=value)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
            cell.fill = header_fill

    print(f"{sheet_name} 工作表已创建，共 {len(df)} 条记录")


def main():
    print("正在获取Upbit交易所的所有交易对...")
    markets = get_upbit_markets()
    if not markets:
        print("无法获取市场信息，程序退出。")
        return

    # 获取上币日期数据（示例实现）
    listing_dates = get_coin_listing_dates()

    # 筛选并排序交易对
    krw_pairs = sort_by_listing_date(filter_pairs(markets, 'KRW-'), listing_dates)
    usdt_pairs = sort_by_listing_date(filter_pairs(markets, 'USDT-'), listing_dates)
    btc_pairs = sort_by_listing_date(filter_pairs(markets, 'BTC-'), listing_dates)

    # 提取交易对名称
    krw_markets = {pair['market'].split('-')[1] for pair in krw_pairs}
    usdt_markets = {pair['market'].split('-')[1] for pair in usdt_pairs}
    btc_markets = {pair['market'].split('-')[1] for pair in btc_pairs}

    # 找出仅在特定市场的交易对并排序
    only_krw = sort_by_listing_date(
        [pair for pair in krw_pairs if
         pair['market'].split('-')[1] not in usdt_markets and pair['market'].split('-')[1] not in btc_markets],
        listing_dates
    )
    only_usdt = sort_by_listing_date(
        [pair for pair in usdt_pairs if
         pair['market'].split('-')[1] not in krw_markets and pair['market'].split('-')[1] not in btc_markets],
        listing_dates
    )
    only_btc = sort_by_listing_date(
        [pair for pair in btc_pairs if
         pair['market'].split('-')[1] not in krw_markets and pair['market'].split('-')[1] not in usdt_markets],
        listing_dates
    )

    # 找出三种市场全有的交易对并排序
    all_markets = sort_by_listing_date(
        [pair for pair in krw_pairs if
         pair['market'].split('-')[1] in usdt_markets and pair['market'].split('-')[1] in btc_markets],
        listing_dates
    )

    # 找出在USDT和BTC市场同时存在，并且不在KRW市场的交易对
    usdt_btc_not_krw = sort_by_listing_date(
        [pair for pair in usdt_pairs if
         pair['market'].split('-')[1] in btc_markets and pair['market'].split('-')[1] not in krw_markets],
        listing_dates
    )

    # 创建output文件夹
    output_dir = 'output'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = os.path.join(output_dir, f'upbit_pairs_{timestamp}.xlsx')

    # 写入Excel文件
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 写入主要市场数据
        save_to_excel(krw_pairs, 'KRW_pairs', writer)
        save_to_excel(usdt_pairs, 'USDT_pairs', writer)
        save_to_excel(btc_pairs, 'BTC_pairs', writer)

        # 写入仅存在于单一市场的数据（确保即使为空也创建sheet）
        save_to_excel(only_krw, 'only_KRW_pairs', writer, 'KRW')
        save_to_excel(only_usdt, 'only_USDT_pairs', writer, 'USDT')
        save_to_excel(only_btc, 'only_BTC_pairs', writer, 'BTC')

        # 写入三种市场全有的数据
        save_to_excel(all_markets, 'all_markets_pairs', writer)

        # 写入在USDT和BTC市场同时存在，并且不在KRW市场的交易对
        save_to_excel(usdt_btc_not_krw, 'usdt_btc_not_krw_pairs', writer)

    print(f"\n数据已成功保存到: {filename}")
    print(
        "包含的工作表有: KRW_pairs, USDT_pairs, BTC_pairs, only_KRW_pairs, only_USDT_pairs, only_BTC_pairs, all_markets_pairs, usdt_btc_not_krw_pairs")
    print("所有工作表均按上币日期（或近似顺序）排序")


if __name__ == "__main__":
    main()