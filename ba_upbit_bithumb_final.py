import requests
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os


class CryptoExchangeAnalyzer:
    def __init__(self):
        self.output_dir = "output"
        os.makedirs(self.output_dir, exist_ok=True)
        self.upbit_markets = None
        self.listing_dates = {}  # 存储上币日期数据

    def get_binance_usdt_pairs(self):
        """获取币安USDT交易对"""
        url = "https://api.binance.com/api/v3/exchangeInfo"
        try:
            print("获取币安USDT交易对...")
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            data = response.json()

            usdt_pairs = []
            for symbol in data['symbols']:
                if (symbol['quoteAsset'] == 'USDT'
                        and symbol['status'] == 'TRADING'
                        and symbol['isSpotTradingAllowed']):
                    price_filter = next((f for f in symbol['filters'] if f['filterType'] == 'PRICE_FILTER'), {})
                    lot_size_filter = next((f for f in symbol['filters'] if f['filterType'] == 'LOT_SIZE'), {})

                    pair_info = {
                        'Symbol': symbol['symbol'],
                        'Base Asset': symbol['baseAsset'],
                        'Quote Asset': symbol['quoteAsset'],
                        'Price Precision': price_filter.get('tickSize', 'N/A'),
                        'Min Qty': lot_size_filter.get('minQty', 'N/A'),
                        'Qty Precision': lot_size_filter.get('stepSize', 'N/A'),
                        'Listing Date': datetime.fromtimestamp(symbol['onboardDate'] / 1000).strftime('%Y-%m-%d')
                        if 'onboardDate' in symbol else 'N/A'
                    }
                    usdt_pairs.append(pair_info)
                    # 保存上币日期
                    if 'onboardDate' in symbol:
                        self.listing_dates[symbol['baseAsset']] = datetime.fromtimestamp(
                            symbol['onboardDate'] / 1000).strftime('%Y-%m-%d')

            print(f"找到 {len(usdt_pairs)} 个币安USDT交易对")
            return pd.DataFrame(usdt_pairs)
        except Exception as e:
            print(f"获取币安数据失败: {e}")
            return pd.DataFrame()

    def get_bithumb_krw_pairs(self):
        """获取Bithumb KRW交易对"""
        try:
            print("获取Bithumb KRW交易对...")
            url = "https://api.bithumb.com/public/ticker/ALL_KRW"
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            data = response.json()

            krw_pairs = []
            for currency in data['data']:
                if currency == 'date':
                    continue
                krw_pairs.append({
                    'Market': f'KRW-{currency}',
                    'Currency': currency,
                    'Korean Name': '',
                    'English Name': ''
                })

            print(f"找到 {len(krw_pairs)} 个Bithumb KRW交易对")
            return pd.DataFrame(krw_pairs)
        except Exception as e:
            print(f"获取Bithumb数据失败: {e}")
            return pd.DataFrame()

    def get_upbit_markets(self):
        """获取Upbit交易所的所有市场信息"""
        if self.upbit_markets is None:
            url = "https://api.upbit.com/v1/market/all"
            headers = {"accept": "application/json", 'User-Agent': 'Mozilla/5.0'}
            try:
                print("获取Upbit市场信息...")
                response = requests.get(url, headers=headers, timeout=10)
                response.raise_for_status()
                self.upbit_markets = response.json()
                # 提取上币日期（Upbit API不直接提供，这里使用模拟值）
                for market in self.upbit_markets:
                    base_currency = market['market'].split('-')[1]
                    if base_currency not in self.listing_dates:
                        self.listing_dates[base_currency] = "未知"
            except requests.exceptions.RequestException as e:
                print(f"获取Upbit数据失败: {e}")
                self.upbit_markets = []
        return self.upbit_markets

    def filter_pairs(self, prefix):
        """从市场信息中筛选出特定前缀的交易对"""
        markets = self.get_upbit_markets()
        return [market for market in markets if market['market'].startswith(prefix)]

    def sort_by_listing_date(self, pairs):
        """按上币日期对交易对进行排序"""
        # 使用上币日期排序，未知日期的交易对放在最后
        return sorted(pairs,
                      key=lambda x: self.listing_dates.get(x['market'].split('-')[1], '9999-99-99'))

    def save_to_excel(self, pairs, sheet_name, writer, base_currency=None):
        """将交易对信息保存到Excel的指定工作表中"""
        workbook = writer.book

        # 如果没有数据，创建空的DataFrame
        if not pairs:
            df = pd.DataFrame(columns=[
                '交易对代码', '基础货币', '报价货币',
                '韩文名称', '英文名称', '市场警告', '上币日期'
            ])
            if base_currency:
                df['基础货币'] = base_currency
        else:
            df = pd.DataFrame(pairs)
            df = df.rename(columns={
                'market': '交易对代码',
                'korean_name': '韩文名称',
                'english_name': '英文名称'
            })

            # 添加基础货币和报价货币列
            if 'only_' in sheet_name:
                base_currency = sheet_name.split('_')[1]
                df['基础货币'] = base_currency
            else:
                df['基础货币'] = sheet_name.split('_')[0].upper()

            df['报价货币'] = df['交易对代码'].apply(lambda x: x.split('-')[1])

            # 添加上币日期列
            df['上币日期'] = df['报价货币'].apply(lambda x: self.listing_dates.get(x, '未知'))

        # 确保所有列存在
        required_columns = ['交易对代码', '基础货币', '报价货币', '韩文名称', '英文名称', '上币日期']
        if 'market_warning' in df.columns:
            required_columns.append('市场警告')
            df = df.rename(columns={'market_warning': '市场警告'})
        else:
            df['市场警告'] = ''
            required_columns.append('市场警告')

        df = df[required_columns]

        # 写入Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = workbook[sheet_name]

        # 设置列宽
        for i, col in enumerate(df.columns):
            column_width = max(len(str(x)) for x in df[col]) if not df.empty else len(col)
            column_width = max(column_width, len(col)) + 2
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = column_width

        # 设置表头样式
        if not df.empty:
            header_font = Font(bold=True)
            header_alignment = Alignment(vertical='top', wrap_text=True)
            header_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')

            for col_num, value in enumerate(df.columns.values):
                cell = worksheet.cell(row=1, column=col_num + 1, value=value)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                cell.fill = header_fill

        print(f"{sheet_name} 工作表已创建，共 {len(df)} 条记录")

    def analyze_exchanges(self):
        """执行所有分析并生成综合Excel报告"""
        print("=== 加密货币交易所数据分析工具 ===")

        # 获取各交易所数据
        binance_df = self.get_binance_usdt_pairs()
        bithumb_df = self.get_bithumb_krw_pairs()
        upbit_markets = self.get_upbit_markets()

        # 处理币安数据
        if not binance_df.empty:
            binance_assets = set(binance_df['Base Asset'])
        else:
            binance_assets = set()

        # 处理Bithumb数据
        if not bithumb_df.empty:
            bithumb_assets = set(bithumb_df['Currency'])
        else:
            bithumb_assets = set()

        # 处理Upbit数据并筛选不同类型交易对
        krw_pairs = self.sort_by_listing_date(self.filter_pairs('KRW-'))
        usdt_pairs = self.sort_by_listing_date(self.filter_pairs('USDT-'))
        btc_pairs = self.sort_by_listing_date(self.filter_pairs('BTC-'))

        # 提取Upbit交易对基础货币 - 修复：定义upbit_assets
        upbit_assets = {pair['market'].split('-')[1] for pair in upbit_markets}
        krw_assets = {pair['market'].split('-')[1] for pair in krw_pairs}
        usdt_assets = {pair['market'].split('-')[1] for pair in usdt_pairs}
        btc_assets = {pair['market'].split('-')[1] for pair in btc_pairs}

        # 计算Upbit特定交易对组合
        only_krw = self.sort_by_listing_date(
            [pair for pair in krw_pairs if
             pair['market'].split('-')[1] not in usdt_assets and pair['market'].split('-')[1] not in btc_assets]
        )
        only_usdt = self.sort_by_listing_date(
            [pair for pair in usdt_pairs if
             pair['market'].split('-')[1] not in krw_assets and pair['market'].split('-')[1] not in btc_assets]
        )
        only_btc = self.sort_by_listing_date(
            [pair for pair in btc_pairs if
             pair['market'].split('-')[1] not in krw_assets and pair['market'].split('-')[1] not in usdt_assets]
        )

        all_upbit_markets = self.sort_by_listing_date(
            [pair for pair in krw_pairs if
             pair['market'].split('-')[1] in usdt_assets and pair['market'].split('-')[1] in btc_assets]
        )

        usdt_btc_not_krw = self.sort_by_listing_date(
            [pair for pair in usdt_pairs if
             pair['market'].split('-')[1] in btc_assets and pair['market'].split('-')[1] not in krw_assets]
        )

        # 计算交易所间的交易对组合
        all_three_exchanges = binance_assets & upbit_assets & bithumb_assets
        only_binance = binance_assets - upbit_assets - bithumb_assets
        only_upbit = upbit_assets - binance_assets - bithumb_assets
        only_bithumb = bithumb_assets - binance_assets - upbit_assets
        binance_upbit = binance_assets & upbit_assets - bithumb_assets
        binance_bithumb = binance_assets & bithumb_assets - upbit_assets
        bithumb_upbit = bithumb_assets & upbit_assets - binance_assets


        # 计算ba_bithumb与usdt_btc_not_krw的比较结果
        # ba_bithumb_assets = binance_bithumb
        binance_bithumb_not_upbit_krw = binance_assets & bithumb_assets - krw_assets
        usdt_btc_not_krw_assets = {pair['market'].split('-')[1] for pair in usdt_btc_not_krw}

        common_pairs = sorted(binance_bithumb_not_upbit_krw & usdt_btc_not_krw_assets)
        only_in_ba_bithumb = sorted(binance_bithumb_not_upbit_krw - usdt_btc_not_krw_assets)
        only_in_upbit = sorted(usdt_btc_not_krw_assets - binance_bithumb_not_upbit_krw)

        # 生成输出文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.output_dir, f"Crypto_Exchange_Analysis_{timestamp}.xlsx")

        # 写入Excel文件
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 写入Upbit各类型交易对
            self.save_to_excel(krw_pairs, 'Upbit_KRW_pairs', writer)
            self.save_to_excel(usdt_pairs, 'Upbit_USDT_pairs', writer)
            self.save_to_excel(btc_pairs, 'Upbit_BTC_pairs', writer)
            self.save_to_excel(only_krw, 'Upbit_only_KRW', writer, 'KRW')
            self.save_to_excel(only_usdt, 'Upbit_only_USDT', writer, 'USDT')
            self.save_to_excel(only_btc, 'Upbit_only_BTC', writer, 'BTC')
            self.save_to_excel(all_upbit_markets, 'Upbit_all_markets', writer)
            self.save_to_excel(usdt_btc_not_krw, 'Upbit_USDT_BTC_not_KRW', writer)

            # 写入交易所间比较结果
            pd.DataFrame(all_three_exchanges, columns=['Asset']).to_excel(writer, sheet_name='All_Exchanges',
                                                                          index=False)
            pd.DataFrame(only_binance, columns=['Asset']).to_excel(writer, sheet_name='Only_Binance', index=False)
            pd.DataFrame(only_upbit, columns=['Asset']).to_excel(writer, sheet_name='Only_Upbit', index=False)
            pd.DataFrame(only_bithumb, columns=['Asset']).to_excel(writer, sheet_name='Only_Bithumb', index=False)
            pd.DataFrame(binance_upbit, columns=['Asset']).to_excel(writer, sheet_name='Binance_Upbit', index=False)
            pd.DataFrame(binance_bithumb, columns=['Asset']).to_excel(writer, sheet_name='Binance_Bithumb', index=False)
            pd.DataFrame(bithumb_upbit, columns=['Asset']).to_excel(writer, sheet_name='Bithumb_Upbit', index=False)

            # 写入ba_bithumb与usdt_btc_not_krw的比较结果
            pd.DataFrame(common_pairs, columns=["Common Pairs"]).to_excel(writer, sheet_name="Common_Pairs",
                                                                          index=False)
            pd.DataFrame(only_in_ba_bithumb, columns=["Only in Binance_Bithumb"]).to_excel(writer,
                                                                                           sheet_name="Only_Binance_Bithumb",
                                                                                           index=False)
            pd.DataFrame(only_in_upbit, columns=["Only in Upbit_USDT_BTC"]).to_excel(writer,
                                                                                     sheet_name="Only_Upbit_USDT_BTC",
                                                                                     index=False)

        print(f"\n所有分析数据已保存到: {filename}")
        print("程序执行完毕！")


if __name__ == "__main__":
    analyzer = CryptoExchangeAnalyzer()
    analyzer.analyze_exchanges()